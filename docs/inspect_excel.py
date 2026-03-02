import openpyxl
import json
import os

def read_excel(file_path):
    if not os.path.exists(file_path):
        return f"File {file_path} not found"
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    data = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_data = []
        for row in sheet.iter_rows(values_only=True):
            sheet_data.append(list(row))
        data[sheet_name] = sheet_data
    return data

docs_dir = "/home/ooka/BACKUP ARCH/jinx/Kerja/CPH-Dashboard/CPH-Dashboard/docs"
file1 = os.path.join(docs_dir, "KUESIONER ONBOARDING CUSTOMER FINAL.xlsx")
file2 = os.path.join(docs_dir, "Customer Onboarding Checklist EXPANDED.xlsx")

print("--- FILE 1: KUESIONER ONBOARDING CUSTOMER FINAL.xlsx ---")
content1 = read_excel(file1)
print(json.dumps(content1, indent=2))

print("\n--- FILE 2: Customer Onboarding Checklist EXPANDED.xlsx ---")
content2 = read_excel(file2)
print(json.dumps(content2, indent=2))

#!/usr/bin/env python3
"""
Generate Excel Checklist V2 — Kesiapan Penggunaan Sistem CPH Tyre Dashboard
========================================================================
Updated: 02 Maret 2026
Ref: Masukan Pak Roland CPH & Onboarding Questionnaire
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
import os

# ============================================================
# COLOR PALETTE
# ============================================================
CPH_RED = "C0392B"
CPH_DARK_RED = "922B21"
CPH_ORANGE = "E67E22"
CPH_DARK = "2C3E50"
CPH_LIGHT_GRAY = "ECF0F1"
CPH_WHITE = "FFFFFF"
CPH_GREEN = "27AE60"
CPH_BLUE = "2980B9"

# ============================================================
# STYLE HELPERS
# ============================================================
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

medium_border = Border(
    left=Side(style='medium'),
    right=Side(style='medium'),
    top=Side(style='medium'),
    bottom=Side(style='medium')
)

def style_title(ws, row, col, text, merge_end_col=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name='Calibri', size=16, bold=True, color=CPH_WHITE)
    cell.fill = PatternFill(start_color=CPH_RED, end_color=CPH_RED, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end_col)
        for c in range(col, merge_end_col + 1):
            ws.cell(row=row, column=c).fill = PatternFill(start_color=CPH_RED, end_color=CPH_RED, fill_type='solid')
            ws.cell(row=row, column=c).border = medium_border

def style_subtitle(ws, row, col, text, merge_end_col=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name='Calibri', size=11, italic=True, color=CPH_DARK)
    cell.fill = PatternFill(start_color=CPH_LIGHT_GRAY, end_color=CPH_LIGHT_GRAY, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end_col)

def style_section_header(ws, row, col, text, merge_end_col=None, color=CPH_DARK):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name='Calibri', size=12, bold=True, color=CPH_WHITE)
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end_col)
        for c in range(col, merge_end_col + 1):
            ws.cell(row=row, column=c).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            ws.cell(row=row, column=c).border = thin_border

def style_header_row(ws, row, headers, start_col=1, color=CPH_DARK_RED):
    for i, header in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=header)
        cell.font = Font(name='Calibri', size=10, bold=True, color=CPH_WHITE)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def style_data_row(ws, row, data, start_col=1, alt=False):
    bg = CPH_LIGHT_GRAY if alt else CPH_WHITE
    for i, val in enumerate(data):
        cell = ws.cell(row=row, column=start_col + i, value=val)
        cell.font = Font(name='Calibri', size=10)
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = thin_border

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# SHEET 1: ROLE & PIC — INTERNAL CPH
# ============================================================
def create_sheet_internal(wb):
    ws = wb.active
    ws.title = "1. PIC Internal CPH"
    ws.sheet_properties.tabColor = CPH_RED

    set_col_widths(ws, [5, 20, 30, 20, 15, 25, 15, 20])
    max_col = 8

    style_title(ws, 1, 1, "CHECKLIST ROLE & PIC — INTERNAL CPH", max_col)
    ws.row_dimensions[1].height = 40
    style_subtitle(ws, 2, 1, "Daftar penanggung jawab internal CPH untuk implementasi sistem", max_col)

    row = 4
    headers = ["No", "Role", "Hak Akses Utama", "Nama PIC", "Title", "Email", "WhatsApp", "Status"]
    style_header_row(ws, row, headers)

    roles = [
        ["1", "Super Admin", "Management user, role, audit log, full control", "", "IT Manager", "", "", "☐ Siap"],
        ["2", "Manajerial", "All dashboard, export, view all reports", "", "Project Manager", "", "", "☐ Siap"],
        ["3", "Supervisor", "Edit & approve data, import management", "", "Supervisor", "", "", "☐ Siap"],
        ["4", "Admin / Operator", "Daily data entry (Tyre/Vehicle/Movement)", "", "Admin", "", "", "☐ Siap"],
        ["5", "IT Support", "Technical troubleshooting, DB coordination", "", "IT Support", "", "", "☐ Siap"],
    ]

    for i, role_data in enumerate(roles):
        r = row + 1 + i
        style_data_row(ws, r, role_data, alt=(i % 2 == 1))
        ws.row_dimensions[r].height = 50

# ============================================================
# SHEET 2: ROLE & PIC — EKSTERNAL (CUSTOMER)
# ============================================================
def create_sheet_external(wb):
    ws = wb.create_sheet("2. PIC Eksternal (Customer)")
    ws.sheet_properties.tabColor = CPH_ORANGE

    # More detail columns as requested by Pak Roland
    set_col_widths(ws, [5, 20, 25, 20, 15, 25, 15, 30, 10])
    max_col = 9

    style_title(ws, 1, 1, "CHECKLIST ROLE & PIC — EKSTERNAL (CUSTOMER)", max_col)
    ws.row_dimensions[1].height = 40
    style_subtitle(ws, 2, 1, "Daftar personil customer yang akan mengelola & menggunakan sistem", max_col)

    row = 4
    headers = ["No", "Role Customer", "Tanggung Jawab Utama", "Nama PIC", "Title", "Email", "WhatsApp", "Kewajiban Data", "Status"]
    style_header_row(ws, row, headers, color=CPH_ORANGE)

    roles = [
        ["1", "PIC Utama / Fleet Mgr", "Penanggung jawab onboarding & validasi data utama", "", "Fleet Manager", "", "", "Validasi Master Data (Vehicle/Tyre)", "☐ Siap"],
        ["2", "Supervisor Fleet", "Monitoring pergerakan ban & jadwal inspeksi", "", "Supervisor", "", "", "Review Report Bulanan", "☐ Siap"],
        ["3", "Admin / Operator", "Input data harian (pemasangan & pelepasan)", "", "Admin", "", "", "Update HM/KM & Serial Number", "☐ Siap"],
        ["4", "Tyreman / Inspector", "Melakukan inspeksi ban fisik di lapangan", "", "Inspector", "", "", "RTD & PSI Measurement Data", "☐ Siap"],
        ["5", "Finance / Viewer", "Melihat laporan CPK & Biaya untuk invoicing", "", "Staff Finance", "", "", "Review Cost Analysis", "☐ Siap"],
    ]

    for i, role_data in enumerate(roles):
        r = row + 1 + i
        style_data_row(ws, r, role_data, alt=(i % 2 == 1))
        ws.row_dimensions[r].height = 60
    
    # Information Section
    info_row = row + len(roles) + 3
    style_section_header(ws, info_row, 1, "💡 CATATAN UNTUK CUSTOMER", max_col, CPH_ORANGE)
    ws.cell(row=info_row+1, column=1, value="1. Pastikan setiap PIC memiliki nomor WhatsApp aktif untuk notifikasi sistem.")
    ws.cell(row=info_row+2, column=1, value="2. Alamat email akan digunakan sebagai Username login.")
    ws.cell(row=info_row+3, column=1, value="3. Penunjukan PIC harus resmi agar proses onboarding berjalan lancar.")

# ============================================================
# SHEET 3: DATA FORMAT & TEMPLATE
# ============================================================
def create_sheet_data_req(wb):
    ws = wb.create_sheet("3. Format Data & Template")
    ws.sheet_properties.tabColor = CPH_BLUE

    set_col_widths(ws, [5, 20, 60, 40, 15])
    max_col = 5

    style_title(ws, 1, 1, "FORMAT DATA & TEMPLATE CSV/EXCEL", max_col)
    ws.row_dimensions[1].height = 40
    style_subtitle(ws, 2, 1, "Detail struktur kolom untuk import data awal ke dalam sistem", max_col)

    row = 4
    headers = ["No", "Modul", "Struktur Kolom (Format CSV/Excel)", "Contoh Data", "Status"]
    style_header_row(ws, row, headers, color=CPH_BLUE)

    data_formats = [
        ["1", "Tyre Master", 
         "serial_number, brand_name, size_name, pattern_name, initial_rtd, location_name, segment_name, price, status",
         "SN-1001, BRIDGESTONE, 11.00-20, G580, 16.5, GUDANG-A, Coal Hauling, 5500000, New", "☐ Siap"],
        
        ["2", "Vehicle Master", 
         "kode_kendaraan, no_polisi, model_kendaraan, brand_kendaraan, site_location, curb_weight, payload_capacity, segment",
         "DT-101, B 1234 ABC, DUMP TRUCK, HINO, SITE-KALTIM, 5500, 20, Coal Hauling", "☐ Siap"],
        
        ["3", "Movement History", 
         "serial_number, kode_kendaraan, movement_type, movement_date, position_code, odometer",
         "SN-1001, DT-101, Installation, 2026-02-28, LF, 50000", "☐ Siap"],
        
        ["4", "Tyre Size", 
         "size, brand_name, type, std_otd, ply_rating",
         "11.00-20, BRIDGESTONE, Radial, 16.5, 16", "☐ Siap"],

        ["5", "Failure Codes", 
         "failure_code, failure_name, default_category",
         "CUT, Cut Separation, Major Damage", "☐ Siap"],
         
        ["6", "Locations", 
         "location_name, location_type, capacity",
         "SITE-KALTIM-GUDANG, Warehouse, 100", "☐ Siap"],
    ]

    for i, item in enumerate(data_formats):
        r = row + 1 + i
        style_data_row(ws, r, item, alt=(i % 2 == 1))
        ws.row_dimensions[r].height = 80
    
    # ---- NEW: TIRE-VEHICLE MAPPING (CRITICAL) ----
    map_row = row + len(data_formats) + 2
    style_section_header(ws, map_row, 1, "📌 MAPPING BAN TERPASANG (CRITICAL FOR ONBOARDING)", max_col, CPH_DARK)
    map_row += 1
    style_header_row(ws, map_row, ["No", "Area / Site", "Unit Code", "Position", "Tyre Serial Number", "Current RTD", "Current PSI", "Odometer Saat Mapping"], color=CPH_DARK)
    
    style_data_row(ws, map_row+1, ["1", "SITE-A", "DT-101", "LF (Left Front)", "SN-99001", "14.5", "110", "45200"])
    style_data_row(ws, map_row+2, ["2", "SITE-A", "DT-101", "RF (Right Front)", "SN-99002", "14.2", "110", "45200"])
    ws.row_dimensions[map_row+1].height = 25
    ws.row_dimensions[map_row+2].height = 25

# ============================================================
# SHEET 4: EXPANDED ONBOARDING CHECKLIST
# ============================================================
def create_sheet_onboarding(wb):
    ws = wb.create_sheet("4. Onboarding Checklist")
    ws.sheet_properties.tabColor = CPH_GREEN

    set_col_widths(ws, [5, 45, 20, 15, 15, 20])
    max_col = 6

    style_title(ws, 1, 1, "EXPANDED ONBOARDING CHECKLIST", max_col)
    ws.row_dimensions[1].height = 40

    row = 3
    headers = ["No", "Aktivitas / Task", "Penanggung Jawab", "Timeline", "Status", "Keterangan"]
    style_header_row(ws, row, headers, color=CPH_GREEN)

    # Phases
    phases = [
        ("PHASE 1: PREPARATION & DATA COLLECTION", [
            ["1.1", "Kick-off meeting & penunjukan PIC utama", "Customer / CPH", "Week 1", "☐", ""],
            ["1.2", "Penyerahan Kuesioner Onboarding (Sheet 5)", "Customer", "Week 1", "☐", ""],
            ["1.3", "Pengumpulan Data Master Kendaraan (100% unit)", "Customer", "Week 1", "☐", "Format Sheet 3"],
            ["1.4", "Pengumpulan Data Master Ban (Stok & Pasang)", "Customer", "Week 1", "☐", "Format Sheet 3"],
            ["1.5", "Pemetaan ban terpasang per unit + RTD aktual", "Customer", "Week 1", "☐", "PENTING"],
        ]),
        ("PHASE 2: SYSTEM CONFIGURATION", [
            ["2.1", "Setup company profile & site location", "CPH", "Week 2", "☐", ""],
            ["2.2", "Review & customize Failure Codes per site", "CPH / Customer", "Week 2", "☐", ""],
            ["2.3", "Import master data kendaraan & ban", "CPH", "Week 2", "☐", ""],
            ["2.4", "Konfigurasi Axle Layout per jenis unit", "CPH", "Week 2", "☐", ""],
            ["2.5", "Pembuatan akun user (Internal & Eksternal)", "CPH", "Week 2", "☐", ""],
        ]),
        ("PHASE 3: TRAINING & GO-LIVE", [
            ["3.1", "Training Fleet Manager & Supervisor", "CPH", "Week 3", "☐", "Usage Analysis"],
            ["3.2", "Training Operator / Admin Lapangan", "CPH", "Week 3", "☐", "Data Entry"],
            ["3.3", "Simulasi & UAT (User Acceptance Test)", "Customer", "Week 3", "☐", ""],
            ["3.4", "Soft Launch (Real Data Entry Monitoring)", "Customer / CPH", "Week 4", "☐", ""],
            ["3.5", "Official Go-Live & Handover", "CPH", "Week 4", "☐", ""],
        ])
    ]

    curr_row = row + 1
    for title, items in phases:
        style_section_header(ws, curr_row, 1, title, max_col, CPH_DARK)
        curr_row += 1
        for i, item in enumerate(items):
            style_data_row(ws, curr_row, item, alt=(i % 2 == 1))
            ws.row_dimensions[curr_row].height = 30
            curr_row += 1
        curr_row += 1

# ============================================================
# SHEET 5: KUESIONER ONBOARDING
# ============================================================
def create_sheet_questionnaire(wb):
    ws = wb.create_sheet("5. Kuesioner Onboarding")
    ws.sheet_properties.tabColor = CPH_DARK

    set_col_widths(ws, [5, 30, 50, 20])
    max_col = 4

    style_title(ws, 1, 1, "KUESIONER ONBOARDING CUSTOMER", max_col)
    ws.row_dimensions[1].height = 40

    questions = [
        ("I. PROFIL PERUSAHAAN & SITE", [
            ["Nama Lengkap Perusahaan", "", ""],
            ["Nama Site / Proyek", "", ""],
            ["Alamat Lokasi Site", "", ""],
            ["Waktu Operasional Site", "", "Contoh: 24/7 atau 2 Shift"],
            ["Total Populasi Kendaraan", "", "Unit"],
        ]),
        ("II. KONTAK PERSON IMPLEMENTASI (PIC)", [
            ["Nama Lengkap PIC", "", ""],
            ["Jabatan / Title", "", ""],
            ["Email (Username Login)", "", ""],
            ["Nomor WhatsApp (Aktif)", "", ""],
            ["PIC IT / Support", "", "Nama & No WA"],
        ]),
        ("III. TEKNIS & HARDWARE", [
            ["Metode Input Data", "", "PC / Tablet / Mobile"],
            ["Ketersediaan Jaringan Internet", "", "Stabil / Terbatas / None"],
            ["Hardware Pendukung (Barcode Scanner)", "", "Tersedia / Belum"],
            ["Metode Penandaan Ban", "", "Barcode / Branding / Painting"],
        ]),
        ("IV. MANAJEMEN DATA", [
            ["Brand ban yang paling banyak digunakan", "", ""],
            ["Sistem pencatatan saat ini", "", "Excel / Manual / SAP / FMS"],
            ["Target Go-Live Penggunaan Sistem", "", "Tanggal"],
            ["Frekuensi Update RTD (Inspeksi)", "", "Mingguan / Bulanan / Per KM"],
        ])
    ]

    curr_row = 3
    for title, items in questions:
        style_section_header(ws, curr_row, 1, title, max_col, CPH_BLUE)
        curr_row += 1
        for i, item in enumerate(items):
            ws.cell(row=curr_row, column=2, value=item[0]).font = Font(bold=True)
            ws.cell(row=curr_row, column=3, value=item[1]).border = thin_border
            ws.cell(row=curr_row, column=4, value=item[2])
            ws.row_dimensions[curr_row].height = 25
            curr_row += 1
        curr_row += 1

# ============================================================
# MAIN
# ============================================================
def main():
    wb = openpyxl.Workbook()
    
    create_sheet_internal(wb)
    create_sheet_external(wb)
    create_sheet_data_req(wb)
    create_sheet_onboarding(wb)
    create_sheet_questionnaire(wb)

    for ws in wb.worksheets:
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.sheet_view.showGridLines = False

    output_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(output_dir, "CPH_Checklist_Onboarding_V2.xlsx")
    wb.save(output_path)
    print(f"✅ Checklist V2 berhasil dibuat: {output_path}")

if __name__ == "__main__":
    main()
